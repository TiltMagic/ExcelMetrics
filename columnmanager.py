from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl
import os
import timestamp
import error_logger
import json

with open("config.json", "r") as config_file:
    configs = json.load(config_file)


class ColumnManager:

    def __init__(self, file_location, sheet_name):
        self.file_location = file_location
        self.file_name = list(file_location.split("\\"))[-1]
        self.workbook = self.load_workbook(file_location)
        self.sheet = self.get_sheet(self.workbook, sheet_name)

    def load_workbook(self, file_location):
        """
        Loads workbook object from excel file
        """
        try:
            return openpyxl.load_workbook(file_location)
        except Exception as traceback_error:
            statement = "Problem laoding workbook from {}".format(file_location)
            error_logger.logger(statement, traceback_error)

    def save_doc(self):
        # Change to enable saving to any location *args
        """
        Saves ColumnManager changes to excel workbook
        """
        try:
            self.workbook.save(self.file_location)
        except Exception as traceback_error:
            statement = "Unable to save file {}".format(self.file_name)
            error_logger.logger(statement, traceback_error)

    def get_sheet(self, workbook_name, sheet_name):
        """
        Retrieves sheet object from workbook object
        """
        try:
            return workbook_name[sheet_name]
        except Exception as traceback_error:
            statement = "Can't find sheet {} for {}".format(
                sheet_name, self.file_name)
            error_logger.logger(statement, traceback_error)

    def make_new_column(self, title, column_start=1):
        """
        Create column w/ given title
        """
        column_titles = self.get_column_titles()

        if title in column_titles:
            statement = "Column title {} NOT added- already exists for file {}".format(
                title, self.file_name)
            error_logger.logger(statement)
        else:
            try:
                max_column = self.sheet.max_column
                self.sheet.cell(row=1, column=max_column+column_start).value = title
                error_logger.logger("{} column title added".format(title))
            except Exception as traceback_error:
                statement = "Something went wrong with building column {}".format(title)
                self.logger(statement, traceback_error)

    def value_exists(self, value):
        '''
        Checks if given value exists anywhere in excel file
        '''
        for row in self.sheet.values:
            for data in row:
                if data == value:
                    return True

    def group_by_exists(self):
        '''
        Checks if cell 'A1' has the value of 'Group By'
        Method is meant to check if formatting should occur
        '''
        try:
            return self.sheet['A1'].value == 'Group By'
        except Exception as traceback_error:
            statement = "Problem reading value from cell 'A1'"
            error_logger.logger(statement, traceback_error)

    def get_column_titles(self, row=1):
        """
        Returns list with column title
        """

        max_column = self.sheet.max_column

        try:
            column_titles = [self.sheet.cell(row=row, column=column).value
                             for column in range(1, max_column+1)]
            return column_titles
        except Exception as traceback_error:
            statement = "Problem retrieving column titles"
            error_logger.logger(statement, traceback_error)

    def get_column_titles_with_index(self, row=1):
        """
        Returns dict with column title and column index
        """
        max_column = self.sheet.max_column
        try:
            column_titles = {self.sheet.cell(row=row, column=column).value: column
                             for column in range(1, max_column+1)}
            return column_titles
        except Exception as traceback_error:
            statement = "Problem retrieving column titles"
            error_logger.logger(statement, traceback_error)

    def get_column_index(self, title, row=1):
        """
        Return index number for given column title
        """
        try:
            columns = self.get_column_titles_with_index(row)
            return columns[title]
        except Exception as traceback_error:
            statement = "Problem finding column titled {}".format(title)
            error_logger.logger(statement, traceback_error)

    def get_column_cells(self, title, row=1):
        """
        Returns list of cell objects from given column title
        """
        index = self.get_column_index(title, row)
        try:
            return [cell for cell in list(self.sheet.columns)[index - 1]]
        except Exception as traceback_error:
            statement = "Problem finding cell for column w/ title {}".format(title)
            error_logger.logger(statement, traceback_error)

    def get_column_cell_names(self, title, row=1):
        """
        Return a list of cell names (ex."A32") for given column
        """
        cells = self.get_column_cells(title, row)
        try:
            cell_names = ["{}{}".format(cell.column, cell.row)
                          for cell in cells]
            return cell_names
        except Execption as traceback_error:
            statement = "Problem building cell name from {} column".format(title)
            error_logger.logger(statement, traceback_error)

    def get_column_values(self, title, row=1):
        """
        Returns column values given title
        """
        try:
            cells = self.get_column_cells(title, row)
            values = [cell.value for cell in cells]
            return values
        except Exception as traceback_error:
            statement = "Problem retrieving cells for column {}".format(title)
            error_logger.logger(statement, traceback_error)

    def add_values_to_cell_range(self, cell_range, values, column_index, row):
        for i in range(len(cell_range)):
            try:
                self.sheet.cell(row=row, column=column_index).value = values[i]

            except Exception as traceback_error:
                # statement = "Problem setting cell values"
                error_logger.logger(traceback_error=traceback_error)
            row += 1

    def set_column_values(self, title, values, row=2):
        """
        Sets cell values for given column
        """
        try:
            index = self.get_column_index(title)
            cells_to_set = self.get_column_cells(title)[1:]
        except Exception as traceback_error:
            statement = "Problem with inputs to set_column_values"
            error_logger.logger(statement, traceback_error)

        if self.add_values_to_cell_range(cells_to_set, values, index, row):
            error_logger.logger("Values added to column {}".format(title))

            try:
                for i in range(len(cells_to_set)):
                    self.sheet.cell(row=row, column=index).value = values[i]
                    row += 1
            except Exception as traceback_error:
                statement = "Problem setting cell values for column {}".format(title)
                error_logger.logger(statement, traceback_error)
                # row += 1

    def print_column_values(self, title, row=1):
        """
        Prints out cell values
        """
        cells = self.get_column_cells(title, row)
        for cell in cells:
            print(cell.value)

    def divide_columns_values(self, first_title, second_title, row=1):
        """
        Divides two columns and return list w/ resulting values
        """
        numerator = self.get_column_values(first_title, row=row)
        denominator = self.get_column_values(second_title, row=row)
        result = []
        for i in range(1, len(numerator)):
            try:
                result.append(numerator[i]/denominator[i])
            except (TypeError, ZeroDivisionError) as e:
                # Should the above exception be logged somehow?
                result.append("None")

        return result

    def divide_columns_formula(self, first_title, second_title, row=1):
        """
        Return list of divided cell values from given columns
        """
        try:
            numerator_names = self.get_column_cell_names(first_title, row=row)
            denominator_names = self.get_column_cell_names(second_title, row=row)
            divided_formulas = ["={}/{}".format(numerator_names[i],
                                                denominator_names[i])
                                for i
                                in range(2, len(numerator_names))]

            return divided_formulas

        except Exception as traceback_error:
            statement = "Problem dividing {} values with {} values".format(
                first_title, second_title)
            error_logger.logger(statement, traceback_error)

    def gen_new_metric_column(self, new_title, numer_column, denom_column, with_formulas=False, row=1):
        '''
        Divides existing numerator column values with demoninator column values
        Creates new column with divided values
        '''
        try:
            self.make_new_column(new_title)
            if with_formulas:
                new_values = self.divide_columns_formula(numer_column, denom_column, row=row)
            else:
                new_values = self.divide_columns_values(numer_column, denom_column, row=row)

            self.set_column_values(new_title, new_values, row=row+1)
            # error_logger.logger("Column {} added with values".format(new_title))
        except Exception as traceback_error:
            print(traceback_error)
            statement = "Trouble with generating {} column".format(new_title)
            error_logger.logger(statement, traceback_error)

    def color_column(self, title, color="B3FFB3"):
        """
        read color from config file, keep 'B3FFB3' as default
        """
        cells = self.get_column_cells(title)[:-1]
        for cell in cells:
            cell.fill = PatternFill(start_color=color,
                                    end_color=color,
                                    fill_type="solid")

    def add_validation(self, validation_string):
        '''
        Inserts a new row then adds a data validation drop down list to cell 'A1'.
        'validation_string' parameter contains list of values added to drop down list.
        '''

        try:
            self.sheet.insert_rows(1)
            dv = DataValidation(type="list", formula1=validation_string, allow_blank=True)
            self.sheet.add_data_validation(dv)
            cell = self.sheet["A1"]
            dv.add(cell)
            error_logger.logger('Drop down menu added')
        except Exception as traceback_error:
            statement = "Trouble with adding validation drop-down menu"
            error_logger.logger(statement, traceback_error)

    def get_base_bid_values(self, title, row=2):
        '''
        Return a list of values for systems that are part of the Base Bid
        Relies on seperate 'Bid Item' column with values of ' || Base Bid' for lookup
        '''

        bid_item_values = self.get_column_values("Bid Item", row=row)
        system_values = self.get_column_values(title, row=row)

        results = []

        for value in range(len(bid_item_values)):
            if bid_item_values[value] == " || Base Bid":
                results.append(system_values[value])
        # print(results)
        # for nu in results:
        #     try:
        #         print(self.workbook[nu].value)
        #     except KeyError:
        #         continue
        return results

    def get_jobtype(self):
        try:
            job_type = self.sheet.cell(row=1, column=1).value
            if job_type is None:
                raise TypeError
            return job_type
        except Exception as traceback_error:
            statement = "\nMake sure you select a job type for file {}".format(
                self.file_name.upper())
            error_logger.logger(statement, traceback_error)
